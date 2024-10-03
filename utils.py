import streamlit as st
import pandas as pd
import io
import pandas as pd
import ast
import warnings
from datetime import datetime
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore")
pd.set_option('display.max_rows', 5000)
pd.set_option('display.max_columns', 1000)
import sys
import re
from datetime import datetime, timedelta
from pytz import timezone, all_timezones
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font
pd.set_option('display.max_rows', 5000)
pd.set_option('display.max_columns', 1000)
from openpyxl import load_workbook

def call_straddle_finder(file_loc, filename, sheet_name, spot_price):
    print('**********************')
    print(file_loc)
    print(filename)
    print('**********************')
    spot = spot_price
    print(file_loc)
    df_excel = pd.read_excel(file_loc, sheet_name=sheet_name)
    df = pd.read_csv('{}'.format(filename))
    date_pattern = r'(\d{1,2})_(\d{1,2})_(\d{4})'
    match = re.search(date_pattern, filename)
    if match:
        day, month, year = match.groups()
        extracted_date = f"{day}.{month}.{year}"
        date = extracted_date
    
    date_obj = datetime.strptime(date, "%d.%m.%Y")
    day_of_week = date_obj.strftime("%A")
    
    df = df[[ 'OI','LTP', 'Strike Price',  'LTP.1', 'OI.1']]
    for col in df.columns:
         df[col] = df[col].replace({',': ''}, regex=True).astype(float).astype('int64')
    
    df['CALL_VAR'] = df['OI']*df['LTP']
    df['PUT_VAR'] = df['OI.1']*df['LTP.1']
    df = df.rename(columns={'OI':'CALL_OI', 'LTP':'CALL_LTP', 'LTP.1':'PUT_LTP', 'OI.1':'PUT_OI'})
    
    def convert_cols(l_df, cols):
        for c in cols:
            if c in ['CALL_OI', 'PUT_OI']:
                l_df[c] = round(l_df[c]/10**5,1)
            else:
                l_df[c] = round(l_df[c]/10**7,1)
        return l_df
    
    final_df = convert_cols(df, ['CALL_OI', 'PUT_OI', 'CALL_VAR','PUT_VAR'])
    final_df = final_df.sort_values(['PUT_VAR','CALL_VAR'], ascending=[False,False]).reset_index(drop=True)
    final_df = final_df[(final_df['PUT_VAR']>10)& (final_df['CALL_VAR']>10)]
    if sheet_name.startswith('Bank'):
        final_df = final_df[(final_df['Strike Price'] % 500 == 0)]
    elif sheet_name.startswith('Nifty'):
        final_df = final_df[(final_df['Strike Price'] % 100 == 0)]
    dic = {'Day':[day_of_week], 'Date':[date], 
           'Spot':[int(spot)], 'Strike':[int(final_df.iloc[0]['Strike Price'])],
            'Call VAR': [final_df.iloc[0]['CALL_VAR']] ,
            'Call OI': [final_df.iloc[0]['CALL_OI']],
            'Call Prem': [final_df.iloc[0]['CALL_LTP']],
             'Collective': final_df.iloc[0]['CALL_LTP']+[final_df.iloc[0]['PUT_LTP']] ,
             'Put Prem': [final_df.iloc[0]['PUT_LTP']] ,
             'Put OI': [final_df.iloc[0]['PUT_OI']], 
            'Put VAR': [final_df.iloc[0]['PUT_VAR']], 
            'lower range': [int([final_df.iloc[0]['Strike Price']] - (final_df.iloc[0]['CALL_LTP']+[final_df.iloc[0]['PUT_LTP']]))] , 
            'upper range': [int([final_df.iloc[0]['Strike Price']] + (final_df.iloc[0]['CALL_LTP']+[final_df.iloc[0]['PUT_LTP']]))] 
          }
    
    new = pd.DataFrame(dic)
    df_excel = pd.concat([df_excel, new], ignore_index=True)
    
    book = load_workbook(file_loc)
    with pd.ExcelWriter(file_loc, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_excel.to_excel(writer, sheet_name=sheet_name, index=False)

    return df_excel